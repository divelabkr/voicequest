#!/usr/bin/env python3
"""VoiceQuest 운영 할일 — 월/주/일별 엑셀 생성. 사용자(운영자) 주체 작업.
코드는 Claude가, 결정·배포·운영은 운영자가. 재생성: python scripts/gen-todo-xlsx.py"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEAL, ORANGE = "0F6E56", "D85A30"  # DESIGN.md 토큰
HEAD_FILL = PatternFill("solid", fgColor=TEAL)
HIGH_FILL = PatternFill("solid", fgColor="FBE4DC")
MED_FILL = PatternFill("solid", fgColor="FFF6E5")
ZEBRA = PatternFill("solid", fgColor="F3F0EB")
thin = Side(style="thin", color="DDDDDD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
PRIORITY_FILL = {"高": HIGH_FILL, "中": MED_FILL, "低": None}

# [할 일, 왜(맥락), 우선순위, 예상시간]
SHEETS = {
  "출시 전(1회)": [
    ("GCP Secret Manager에 GEMINI_KEY 추가", "STT 폴백(Deepgram→Gemini) 활성화 — 없으면 Deepgram 단독·무중단", "高", "10분"),
    ("Cloud Run 재배포(--source . --project voicequest-dev)", "Phase0(STT폴백)+복구(atomic write) 코드 클라우드 반영", "高", "20분"),
    ("monitoring 알림 채널 verify(이메일 divelab.kr)", "uptime·에러 알림 수신 — 인증 1회", "高", "5분"),
    ("클라우드 버킷 restore drill 1회", "실제 버킷 versioning 복원 검증(docs/ops/disaster-recovery §4-B)", "中", "20분"),
    ("cloudbuild trigger 연동", "git push 자동배포 — 1회 설정", "中", "30분"),
    ("알파 테스터 모집(초대코드 25명)", "파일럿 게이트 alpha_filled=25 충족", "高", "-"),
    ("GCP 예산 알림 동작 확인(5000KRW)", "비용 폭주 방지 — 알림 실제 수신 확인", "中", "5분"),
  ],
  "일별(Daily)": [
    ("Cloud Run 에러 로그 스캔", "5xx·judge 폴백·STT 폴백 발동 빈도 확인", "高", "5분"),
    ("uptime 확인", "서비스 다운 조기 감지", "高", "2분"),
    ("비용 확인(GCP 콘솔)", "예산 대비 일일 소비 추적", "中", "3분"),
    ("신규 가입·초대 현황(admin 콘솔)", "성장·바이럴 K-factor 체감", "中", "5분"),
    ("사용자 문의·피드백 확인", "이탈 신호·버그 조기 포착", "中", "10분"),
  ],
  "주별(Weekly)": [
    ("restore drill 실행(pnpm drill)", "백업 복구 검증 — '백업 있다≠복구된다'", "高", "5분"),
    ("리텐션 지표 리뷰(D1/D7)", "파일럿 게이트 d1_retention≥40% — 바이럴 전제", "高", "20분"),
    ("파일럿 게이트 점검", "want_replay≥4·alpha_filled 25 — 다음 단계 조건", "高", "15분"),
    ("주간 비용 추세 리뷰", "스케일·번레이트 판단", "中", "10분"),
    ("버킷 versioning 백업 확인", "백업 실제 쌓이는지 — 무음 실패 방지", "中", "5분"),
    ("콘텐츠 진행(에피소드·음성 캐시)", "트레드밀 완화 — 신규 떡밥 공급", "中", "-"),
  ],
  "월별(Monthly)": [
    ("Phase 1(상태 외부화) 착수 결정", "트래픽(동시접속·재시작) 기반 — 스케일 전제, Firestore PITR 동반", "高", "-"),
    ("법률 검토(변호사)", "정식 출시 전 약관·개인정보·국외이전 동의", "高", "-"),
    ("연령 게이트 UI·SynthID 라벨·IAP 적용", "정식 출시 전 컴플라이언스", "高", "-"),
    ("인프라 스케일 검토(max/min-instances)", "바이럴 성공=장애 역설 해소 — Phase 2", "中", "-"),
    ("마케팅 캠페인 점검(덕질 채널)", "바이럴 K≥1.5 레버 — 공유·초대 전환율", "中", "-"),
    ("보안 점검(ADMIN_TOKEN·secret 로테이션)", "키 노출·brute-force 방어 유지", "中", "30분"),
    ("멀티리전 DR 검토(Phase 3)", "스케일 시 리전 장애 대비 — 서울+도쿄", "低", "-"),
  ],
}
HEADERS = ["#", "할 일", "왜 (맥락)", "우선순위", "예상시간", "완료"]
WIDTHS = [4, 42, 54, 10, 10, 8]

wb = Workbook(); wb.remove(wb.active)

cover = wb.create_sheet("📋 개요")
cover.sheet_view.showGridLines = False
cover["B2"] = "VoiceQuest 운영 할일"; cover["B2"].font = Font(size=20, bold=True, color=TEAL)
cover["B3"] = "월·주·일별 운영자 체크리스트 (생성 2026-06-26)"; cover["B3"].font = Font(size=11, color="888888")
intro = [
  "", "이 파일은 '내가(운영자) 해야 할 일'입니다 — 코드는 Claude가, 결정·배포·운영은 운영자가.", "",
  "시트 구성:",
  "   • 출시 전(1회) — 알파를 띄우기 위해 지금 해야 할 일회성 작업",
  "   • 일별(Daily) — 출시 후 매일 5~20분 운영 루틴",
  "   • 주별(Weekly) — 지표·백업·콘텐츠 점검",
  "   • 월별(Monthly) — Phase 진행·법률·마케팅 마일스톤", "",
  "우선순위: 高(빨강) 中(노랑) 低 — '완료' 칸에 체크하며 사용하세요.", "",
  "근거: DR 2축(가용성=STT폴백 PR#14 / 내구성=atomic write+drill PR#15) 완료.",
  "      Phase1~3·법률·IAP는 알파 후/정식 출시 전.",
]
for i, line in enumerate(intro, start=5):
  cover.cell(row=i, column=2, value=line).font = Font(size=11, color="333333")
cover.column_dimensions["A"].width = 2; cover.column_dimensions["B"].width = 92

for name, rows in SHEETS.items():
  ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
  for col, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF", size=11); c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDER
    ws.column_dimensions[get_column_letter(col)].width = w
  ws.row_dimensions[1].height = 24; ws.freeze_panes = "A2"
  for idx, (task, why, pri, dur) in enumerate(rows, start=1):
    r = idx + 1
    for col, v in enumerate([idx, task, why, pri, dur, "☐"], start=1):
      c = ws.cell(row=r, column=col, value=v); c.border = BORDER
      c.alignment = Alignment(vertical="center", wrap_text=(col in (2, 3)),
                              horizontal="center" if col in (1, 4, 5, 6) else "left")
      c.font = Font(size=10)
      if idx % 2 == 0: c.fill = ZEBRA
    pf = PRIORITY_FILL.get(pri)
    if pf:
      pc = ws.cell(row=r, column=4); pc.fill = pf
      pc.font = Font(size=10, bold=True, color=(ORANGE if pri == "高" else "B8860B"))
    ws.row_dimensions[r].height = 30

out = "VoiceQuest_운영_할일.xlsx"
wb.save(out); print("saved:", out)
